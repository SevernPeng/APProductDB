import re
from collections import Counter
from pathlib import Path

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction
from openpyxl import load_workbook

from catalog.models import Brand, Category, Product, ProductModel, ProductType
from catalog.product_types import (
    PRODUCT_TYPE_DEFINITIONS,
    infer_product_type_code,
    product_type_code,
)
from catalog.regions import OFFICIAL_LATEST_HARDWARE, canonical_product_region
from catalog.services import select_comparison_template
from comparison.models import BenchmarkCase, ProductMatch

SHEET_CONFIG = {
    "Access Point": {
        "category": ("Wireless", "Access Point", "access-point"),
        "form_factor": "infer",
    },
    "Gateway": {
        "category": ("Routing", "Gateway", "gateway"),
        "form_factor": Product.APType.OTHER,
    },
    "Bridge": {
        "category": ("Wireless", "Wireless Bridge", "wireless-bridge"),
        "form_factor": Product.APType.OUTDOOR,
    },
    "Managed Switch": {
        "category": ("Switching", "Managed Switch", "managed-switches"),
        "form_factor": Product.APType.OTHER,
    },
    "Unmanaged_EasySmart Switch": {
        "category": (
            "Switching",
            "Unmanaged / Easy Smart Switch",
            "unmanaged-easy-smart-switches",
        ),
        "form_factor": Product.APType.OTHER,
    },
    "Accessories": {
        "category": (None, "Accessories", "accessories"),
        "form_factor": Product.APType.OTHER,
    },
}

BRAND_ALIASES = {
    "tp-link": "TP-Link",
    "tp-link ap 型号": "TP-Link",
    "ubituiti": "Ubiquiti",
    "ubiquiti": "Ubiquiti",
    "uisp": "Ubiquiti",
    "unifi": "Ubiquiti",
    "ruijie / reyee": "Ruijie / Reyee",
    "ruijie/reyee": "Ruijie / Reyee",
    "cisco/meraki": "Cisco / Meraki",
    "meraki / cisco": "Cisco / Meraki",
    "ruckus": "RUCKUS",
    "netgear": "NETGEAR",
    "grandstream": "Grandstream",
    "hikvision": "Hikvision",
    "engenius": "EnGenius",
    "fortinet": "Fortinet",
    "trendnet": "TRENDnet",
    "mikrotik": "MikroTik",
}

MULTI_MODEL_PREFIXES = (
    "RG-", "DS-", "DGS-", "DXS-", "DBA-", "DAP-", "AirEngine ",
    "eKit ", "ECW", "EWS", "FAP-", "GWN", "MR", "CW", "WAX", "NWA",
    "USW-", "UAP-", "U6-", "U7-", "AP-", "ICX", "TPE-", "TEG-",
    "DH-", "WBC", "PFM", "Building Bridge", "Device Bridge",
)


def clean_text(value):
    if value is None:
        return ""
    return str(value).replace("\u00a0", " ").strip()


def split_models(value):
    text = clean_text(value)
    if not text or text in {"-", "–", "—"}:
        return []
    candidates = []
    for line in re.split(r"[\r\n]+", text):
        line = line.strip(" \t•;，,")
        if not line:
            continue
        slash_parts = re.split(r"\s*/\s*", line)
        # Slashes commonly belong to vendor suffixes such as /M(C), /ME,
        # /EI, or /O. Split only when every following part starts like a
        # complete model from a known product family.
        if len(slash_parts) > 1 and all(
            any(
                part.casefold().startswith(prefix.casefold())
                for prefix in MULTI_MODEL_PREFIXES
            )
            for part in slash_parts[1:]
        ):
            candidates.extend(part.strip() for part in slash_parts if part.strip())
        else:
            candidates.append(line)

    expanded = []
    prefix_pattern = "|".join(re.escape(prefix) for prefix in MULTI_MODEL_PREFIXES)
    repeated_prefix = re.compile(rf"\s+(?=(?:{prefix_pattern}))", re.IGNORECASE)
    for candidate in candidates:
        expanded.extend(piece.strip() for piece in repeated_prefix.split(candidate) if piece.strip())

    unique = []
    seen = set()
    for candidate in expanded:
        key = candidate.casefold()
        if key not in seen:
            seen.add(key)
            unique.append(candidate)
    return unique


def parse_model(value):
    model = clean_text(value)
    lifecycle = Product.LifecycleStatus.UNKNOWN
    if re.search(r"\(\s*plan\s+eol\s*\)", model, re.IGNORECASE):
        lifecycle = Product.LifecycleStatus.ANNOUNCED
        model = re.sub(r"\s*\(\s*plan\s+eol\s*\)\s*", "", model, flags=re.IGNORECASE)
    elif re.search(r"\(\s*eol\s*\)", model, re.IGNORECASE):
        lifecycle = Product.LifecycleStatus.DISCONTINUED
        model = re.sub(r"\s*\(\s*eol\s*\)\s*", "", model, flags=re.IGNORECASE)

    hardware_version = ""
    version_match = re.search(r"\s+v(\d[\w.]*)$", model, re.IGNORECASE)
    if version_match:
        hardware_version = f"V{version_match.group(1)}"
        model = model[: version_match.start()].strip()
    return model.strip(), hardware_version, lifecycle


def infer_ap_type(model):
    lowered = model.casefold()
    if "outdoor" in lowered or "bridge" in lowered:
        return Product.APType.OUTDOOR
    if "wall" in lowered or "in-wall" in lowered or lowered.endswith("-iw"):
        return Product.APType.WALL_PLATE
    if "desktop" in lowered:
        return Product.APType.DESKTOP
    if "extender" in lowered:
        return Product.APType.OTHER
    return Product.APType.CEILING


def normalize_brand(header, model):
    normalized_header = clean_text(header).casefold()
    brand_name = BRAND_ALIASES.get(normalized_header, clean_text(header))
    upper_model = model.upper()
    if brand_name == "Ruijie / Reyee":
        if upper_model.startswith(("RG-AP", "RG-SAP", "RG-NBR")):
            return "Ruijie", model
        return "Reyee", model
    if brand_name == "Cisco / Meraki":
        if upper_model.startswith(("MR", "MS")):
            return "Meraki", model
        return "Cisco", model
    if normalized_header == "others":
        vendor_match = re.match(r"^(Mikrotik|MikroTik|Draytek)\s+(.+)$", model, re.IGNORECASE)
        if vendor_match:
            vendor = "MikroTik" if vendor_match.group(1).casefold() == "mikrotik" else "DrayTek"
            return vendor, vendor_match.group(2).strip()
        return "Other", model
    return brand_name, model


def slugify_brand(name):
    value = re.sub(r"[^a-z0-9]+", "-", name.casefold()).strip("-")
    return value or "other"


def upsert_product(
    *,
    brand,
    category,
    model,
    model_key,
    region,
    hardware_version,
    ap_type,
    lifecycle_status,
    source_note,
):
    inferred_type_code = infer_product_type_code(category.slug, model, ap_type)
    product_type = ProductType.objects.filter(
        category=category,
        code=inferred_type_code,
        active=True,
    ).first()
    latest = OFFICIAL_LATEST_HARDWARE.get((brand.slug, model_key, region))
    if latest:
        hardware_version, official_url = latest
    else:
        official_url = ""
    product = (
        Product.objects.filter(brand=brand, model_key=model_key, region=region)
        .order_by("pk")
        .first()
    )
    created = product is None
    if created:
        product = Product.objects.create(
            brand=brand,
            model_key=model_key,
            region=region,
            hardware_version=hardware_version,
            category=category,
            product_type=product_type,
            model=model,
            ap_type=ap_type if category.slug == "access-point" else "",
            wifi_standard="Unknown" if category.slug in {"access-point", "wireless-bridge"} else "",
            lifecycle_status=lifecycle_status,
            notes=source_note,
            is_published=True,
            official_url=official_url,
        )
    if created:
        return product, True

    changed = []
    if (
        product.category_id == category.id
        and product_type
        and product.product_type_id != product_type.pk
    ):
        product.product_type = product_type
        changed.append("product_type")
    if hardware_version and product.hardware_version != hardware_version:
        product.hardware_version = hardware_version
        changed.append("hardware_version")
    if official_url and product.official_url != official_url:
        product.official_url = official_url
        changed.append("official_url")
    if product.lifecycle_status == Product.LifecycleStatus.UNKNOWN and lifecycle_status != Product.LifecycleStatus.UNKNOWN:
        product.lifecycle_status = lifecycle_status
        changed.append("lifecycle_status")
    if not product.is_published:
        product.is_published = True
        changed.append("is_published")
    if source_note not in (product.notes or ""):
        category_note = ""
        if product.category_id != category.id:
            category_note = f"; also mapped under {category}"
        product.notes = f"{product.notes}\n{source_note}{category_note}".strip()
        changed.append("notes")
    if changed:
        product.save(update_fields=tuple(changed) + ("updated_at",))
    return product, False


class Command(BaseCommand):
    help = "Import the cross-brand benchmark matrix workbook into the product database."

    def add_arguments(self, parser):
        parser.add_argument("workbook", type=Path)
        parser.add_argument("--dry-run", action="store_true")
        parser.add_argument(
            "--sheet",
            dest="sheets",
            action="append",
            choices=tuple(SHEET_CONFIG),
            help="Import only the selected sheet. May be repeated.",
        )
        parser.add_argument(
            "--prune-stale",
            action="store_true",
            help=(
                "Treat selected workbook sheets as authoritative and remove existing "
                "benchmark candidates that are absent from them."
            ),
        )
        parser.add_argument(
            "--unpublish-orphans",
            action="store_true",
            help=(
                "Unpublish source-imported competitor products in selected categories "
                "when they have no official URL, specifications, or remaining matches."
            ),
        )

    @transaction.atomic
    def handle(self, *args, **options):
        workbook_path = options["workbook"].resolve()
        if not workbook_path.is_file():
            raise CommandError(f"Workbook not found: {workbook_path}")

        workbook = load_workbook(workbook_path, read_only=True, data_only=True)
        selected_sheets = options["sheets"] or list(SHEET_CONFIG)
        if isinstance(selected_sheets, str):
            selected_sheets = [selected_sheets]
        missing_sheets = set(selected_sheets) - set(workbook.sheetnames)
        if missing_sheets:
            workbook.close()
            raise CommandError(f"Missing required sheets: {', '.join(sorted(missing_sheets))}")

        counts = Counter()
        source_note = f"Imported from {workbook_path.name}"
        selected_category_ids = set()

        for sheet_name in selected_sheets:
            config = SHEET_CONFIG[sheet_name]
            sheet = workbook[sheet_name]
            headers = [clean_text(cell.value) for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
            parent_name, category_name, category_slug = config["category"]
            region = canonical_product_region(category_slug)
            expected_competitors_by_case = {}
            parent = None
            if parent_name:
                parent_slug = re.sub(r"[^a-z0-9]+", "-", parent_name.casefold()).strip("-")
                parent, _ = Category.objects.update_or_create(
                    name=parent_name,
                    defaults={"slug": parent_slug, "parent": None, "active": True},
                )
            category, _ = Category.objects.update_or_create(
                name=category_name,
                defaults={"slug": category_slug, "parent": parent, "active": True},
            )
            for order, (code, name, description) in enumerate(
                PRODUCT_TYPE_DEFINITIONS.get(category_slug, ()),
                start=1,
            ):
                ProductType.objects.update_or_create(
                    category=category,
                    code=code,
                    defaults={
                        "name": name,
                        "description": description,
                        "display_order": order * 10,
                        "active": True,
                    },
                )
            selected_category_ids.add(category.pk)

            for row_number, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                anchor_raw = clean_text(row[0] if row else None)
                if not anchor_raw:
                    continue
                anchor_models = split_models(anchor_raw)
                if len(anchor_models) != 1:
                    counts["ambiguous_anchor_rows"] += 1
                    continue
                anchor_model, anchor_hw, anchor_lifecycle = parse_model(anchor_models[0])
                ap_type = (
                    infer_ap_type(anchor_model)
                    if config["form_factor"] == "infer"
                    else config["form_factor"]
                )
                own_brand, _ = Brand.objects.update_or_create(
                    name="TP-Link",
                    defaults={
                        "slug": "tp-link",
                        "is_own_brand": True,
                        "active": True,
                    },
                )
                anchor, created = upsert_product(
                    brand=own_brand,
                    model_key=re.sub(r"[\s\-_]+", "", anchor_model).upper(),
                    region=region,
                    hardware_version=anchor_hw,
                    category=category,
                    model=anchor_model,
                    ap_type=ap_type,
                    lifecycle_status=anchor_lifecycle,
                    source_note=source_note,
                )
                counts["anchor_products_created" if created else "anchor_products_updated"] += 1

                template = select_comparison_template(category, product_type_code(anchor))
                benchmark_case, case_created = BenchmarkCase.objects.update_or_create(
                    anchor_product=anchor,
                    region=region,
                    name=f"{anchor.model} {region} competitor benchmark",
                    defaults={
                        "template": template,
                        "status": BenchmarkCase.Status.APPROVED,
                        "scenario": sheet_name,
                        "notes": f"{source_note}, sheet {sheet_name}, row {row_number}",
                    },
                )
                counts["benchmark_cases_created" if case_created else "benchmark_cases_updated"] += 1
                expected_competitors = expected_competitors_by_case.setdefault(
                    benchmark_case.pk,
                    set(),
                )

                rank = 0
                for column_index, cell_value in enumerate(row[1:], start=1):
                    header = headers[column_index] if column_index < len(headers) else "Other"
                    for raw_competitor in split_models(cell_value):
                        competitor_model, competitor_hw, competitor_lifecycle = parse_model(raw_competitor)
                        brand_name, competitor_model = normalize_brand(header, competitor_model)
                        if not competitor_model:
                            continue
                        brand, _ = Brand.objects.update_or_create(
                            name=brand_name,
                            defaults={
                                "slug": slugify_brand(brand_name),
                                "is_own_brand": False,
                                "active": True,
                            },
                        )
                        competitor, competitor_created = upsert_product(
                            brand=brand,
                            model_key=re.sub(r"[\s\-_]+", "", competitor_model).upper(),
                            region=region,
                            hardware_version=competitor_hw,
                            category=category,
                            model=competitor_model,
                            ap_type=ap_type,
                            lifecycle_status=competitor_lifecycle,
                            source_note=source_note,
                        )
                        counts[
                            "competitor_products_created"
                            if competitor_created
                            else "competitor_products_updated"
                        ] += 1
                        expected_competitors.add(competitor.pk)
                        rank += 1
                        _, match_created = ProductMatch.objects.update_or_create(
                            benchmark_case=benchmark_case,
                            competitor_product=competitor,
                            defaults={
                                "our_product": anchor,
                                "region": region,
                                "match_type": ProductMatch.MatchType.DIRECT,
                                "match_level": ProductMatch.MatchLevel.CORE,
                                "status": ProductMatch.Status.CONFIRMED,
                                "rank": rank,
                                "reason": f"Mapped in {workbook_path.name}, {sheet_name}!{row_number}",
                            },
                        )
                        counts["matches_created" if match_created else "matches_updated"] += 1
                        if match_created and options["verbosity"] >= 2:
                            self.stdout.write(
                                "CREATE MATCH: "
                                f"{anchor.model} -> {brand.name} {competitor.model} "
                                f"({sheet_name}!{row_number})"
                            )

            if options["prune_stale"]:
                for case_id, expected_competitor_ids in expected_competitors_by_case.items():
                    stale_matches = ProductMatch.objects.filter(benchmark_case_id=case_id)
                    if expected_competitor_ids:
                        stale_matches = stale_matches.exclude(
                            competitor_product_id__in=expected_competitor_ids
                        )
                    if options["verbosity"] >= 2:
                        for stale_match in stale_matches.select_related(
                            "our_product",
                            "competitor_product__brand",
                        ):
                            self.stdout.write(
                                "PRUNE MATCH: "
                                f"{stale_match.our_product.model} -> "
                                f"{stale_match.competitor_product.brand.name} "
                                f"{stale_match.competitor_product.model}"
                            )
                    stale_count, _ = stale_matches.delete()
                    counts["matches_pruned"] += stale_count

        if options["unpublish_orphans"]:
            orphan_products = (
                Product.objects.filter(
                    category_id__in=selected_category_ids,
                    brand__is_own_brand=False,
                    is_published=True,
                    official_url="",
                    notes__contains=source_note,
                    specs__isnull=True,
                    matched_as_competitor__isnull=True,
                )
                .select_related("brand", "category")
                .distinct()
            )
            orphan_product_model_ids = set()
            if options["verbosity"] >= 2:
                for product in orphan_products:
                    self.stdout.write(
                        "UNPUBLISH ORPHAN: "
                        f"{product.brand.name} {product.model} "
                        f"({product.category.name})"
                    )
                    if product.product_model_id:
                        orphan_product_model_ids.add(product.product_model_id)
            else:
                orphan_product_model_ids.update(
                    orphan_products.exclude(product_model_id=None).values_list(
                        "product_model_id",
                        flat=True,
                    )
                )
            orphan_count = orphan_products.update(is_published=False)
            counts["products_unpublished"] += orphan_count
            for product_model_id in orphan_product_model_ids:
                if not Product.objects.filter(
                    product_model_id=product_model_id,
                    is_published=True,
                ).exists():
                    ProductModel.objects.filter(pk=product_model_id).update(active=False)

        workbook.close()
        if options["dry_run"]:
            transaction.set_rollback(True)
        mode = "DRY RUN" if options["dry_run"] else "IMPORTED"
        self.stdout.write(self.style.SUCCESS(f"{mode}: {dict(sorted(counts.items()))}"))
