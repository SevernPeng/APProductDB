from io import BytesIO
from urllib.parse import urlencode

from django.http import HttpResponse, HttpResponseBadRequest
from django.shortcuts import get_object_or_404, render
from django.urls import reverse
from django.views.decorators.http import require_GET
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from catalog.models import Product, normalize_model_key
from core.permissions import catalog_access_required

from .models import ProductMatch
from .services import (
    build_comparison_rows,
    group_comparison_rows,
    load_comparison_products,
    parse_product_ids,
)

BENCHMARK_SEARCH_LIMIT = 20


def _all_published_products():
    return Product.objects.filter(is_published=True).select_related("brand", "category").order_by(
        "category__name", "brand__name", "model", "region", "hardware_version"
    )


def _comparison_query(
    product_ids, only_differences=False, show_sources=False, include_extended=False
):
    parameters = [("products", product_id) for product_id in product_ids]
    if only_differences:
        parameters.append(("differences", "1"))
    if show_sources:
        parameters.append(("sources", "1"))
    if include_extended:
        parameters.append(("extended", "1"))
    return urlencode(parameters)


def _excel_safe(value):
    if isinstance(value, str) and value.startswith(("=", "+", "-", "@")):
        return "'" + value
    return value


@catalog_access_required
@require_GET
def benchmark(request):
    own_products = Product.objects.filter(
        is_published=True, brand__is_own_brand=True
    ).select_related("brand").order_by("model", "region", "hardware_version")
    query = request.GET.get("q", "").strip()
    search_results = []
    search_result_count = 0
    if query:
        normalized_query = normalize_model_key(query)
        if normalized_query:
            matching_products = own_products.filter(
                model_key__icontains=normalized_query
            )
            search_result_count = matching_products.count()
            search_results = list(matching_products[:BENCHMARK_SEARCH_LIMIT])

    selected_id = request.GET.get("product", "").strip()
    if selected_id:
        selected_product = get_object_or_404(own_products, pk=selected_id)
    elif query:
        selected_product = None
    else:
        selected_product = own_products.filter(model_key="EAP772").first()
        if selected_product is None:
            selected_product = own_products.first()

    matches = []
    if selected_product:
        queryset = (
            ProductMatch.objects.filter(
                our_product=selected_product,
                competitor_product__is_published=True,
            )
            .exclude(status=ProductMatch.Status.REJECTED)
            .select_related("competitor_product__brand")
            .order_by("competitor_product__brand__name", "competitor_product__model")
        )
        for match in queryset:
            match.compare_url = (
                reverse("comparison:compare")
                + "?"
                + _comparison_query(
                    [selected_product.pk, match.competitor_product_id]
                )
            )
            matches.append(match)

    compare_all_url = ""
    if matches:
        compare_ids = [selected_product.pk] + [
            match.competitor_product_id for match in matches[:3]
        ]
        compare_all_url = (
            reverse("comparison:compare")
            + "?"
            + _comparison_query(compare_ids)
        )

    return render(
        request,
        "comparison/benchmark.html",
        {
            "query": query,
            "search_results": search_results,
            "search_result_count": search_result_count,
            "search_limit": BENCHMARK_SEARCH_LIMIT,
            "selected_product": selected_product,
            "matches": matches,
            "compare_all_url": compare_all_url,
        },
    )


@catalog_access_required
@require_GET
def compare(request):
    all_products = _all_published_products()
    only_differences = request.GET.get("differences") == "1"
    show_sources = request.GET.get("sources") == "1"
    include_extended = request.GET.get("extended") == "1"
    try:
        product_ids = parse_product_ids(request.GET.getlist("products"))
        if not product_ids:
            return render(
                request,
                "comparison/compare.html",
                {
                    "all_products": all_products,
                    "products": [],
                    "selector_slots": [None, None],
                    "only_differences": only_differences,
                    "show_sources": show_sources,
                    "include_extended": include_extended,
                },
            )
        products = load_comparison_products(product_ids)
    except ValueError as exc:
        return render(
            request,
            "comparison/compare.html",
            {
                "all_products": all_products,
                "products": [],
                "selector_slots": [None, None],
                "error": str(exc),
                "only_differences": only_differences,
                "show_sources": show_sources,
                "include_extended": include_extended,
            },
            status=400,
        )

    rows = build_comparison_rows(
        products,
        only_differences=only_differences,
        include_extended=include_extended,
    )
    for product in products:
        remaining_ids = [item.pk for item in products if item.pk != product.pk]
        product.remove_url = (
            reverse("comparison:compare")
            + "?"
            + _comparison_query(
                remaining_ids, only_differences, show_sources, include_extended
            )
            if len(remaining_ids) >= 2
            else reverse("comparison:benchmark")
        )

    query = _comparison_query(
        product_ids, only_differences, show_sources, include_extended
    )
    export_url = reverse("comparison:export") + "?" + query
    selector_slots = list(products)
    if len(selector_slots) < 4:
        selector_slots.append(None)
    while len(selector_slots) < 2:
        selector_slots.append(None)
    return render(
        request,
        "comparison/compare.html",
        {
            "all_products": all_products,
            "products": products,
            "selector_slots": selector_slots,
            "grouped_rows": group_comparison_rows(rows),
            "row_count": len(rows),
            "only_differences": only_differences,
            "show_sources": show_sources,
            "include_extended": include_extended,
            "export_url": export_url,
            "current_query": query,
        },
    )


@catalog_access_required
@require_GET
def export_comparison(request):
    only_differences = request.GET.get("differences") == "1"
    show_sources = request.GET.get("sources") == "1"
    include_extended = request.GET.get("extended") == "1"
    try:
        product_ids = parse_product_ids(request.GET.getlist("products"))
        products = load_comparison_products(product_ids)
    except ValueError as exc:
        return HttpResponseBadRequest(str(exc))

    rows = build_comparison_rows(
        products,
        only_differences=only_differences,
        include_extended=include_extended,
    )
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Comparison"

    headers = ["Specification"]
    for product in products:
        label = f"{product.brand.name} {product.model}"
        headers.append(label)
        if show_sources:
            headers.append(f"{label} Source")
    worksheet.append([_excel_safe(value) for value in headers])

    for row in rows:
        values = [row["name"]]
        for value in row["values"]:
            values.append(value["display"])
            if show_sources:
                values.append(value["source_url"])
        worksheet.append([_excel_safe(value) for value in values])

    header_fill = PatternFill("solid", fgColor="D9EAF7")
    difference_fill = PatternFill("solid", fgColor="FFF2CC")
    for cell in worksheet[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
    for row_number, row in enumerate(rows, start=2):
        if row["is_different"]:
            for cell in worksheet[row_number]:
                cell.fill = difference_fill
    worksheet.freeze_panes = "B2"
    worksheet.auto_filter.ref = worksheet.dimensions
    worksheet.column_dimensions["A"].width = 28
    for column in worksheet.iter_cols(min_col=2, max_col=worksheet.max_column):
        worksheet.column_dimensions[column[0].column_letter].width = 24

    output = BytesIO()
    workbook.save(output)
    response = HttpResponse(
        output.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = 'attachment; filename="product-comparison.xlsx"'
    return response
