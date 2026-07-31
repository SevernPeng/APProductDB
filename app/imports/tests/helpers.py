from io import BytesIO

from django.core.files.uploadedfile import SimpleUploadedFile
from openpyxl import Workbook

from imports.models import ImportJob
from imports.template_profiles import get_template_profile

SPEC_COLUMNS = [
    "Brand",
    "Model",
    "Region / HW Version",
    "AP Type",
    "Wi-Fi Standard",
    "Supported Wireless Bands",
    "Total Spatial Streams",
    "2.4 GHz MIMO",
    "5 GHz MIMO",
    "6 GHz MIMO",
    "2.4 GHz Max Rate (Mbps)",
    "5 GHz Max Rate (Mbps)",
    "6 GHz Max Rate (Mbps)",
    "Aggregate Rate (Mbps)",
    "Max Channel Width (MHz)",
    "Ethernet Interfaces",
    "PoE Input",
    "PoE Output",
    "Max Clients",
    "IP Rating",
    "Official Source",
    "Last Verified",
    "Data Notes",
]

MATCH_COLUMNS = [
    "TP-Link Model",
    "Competitor Brand 1",
    "Competitor Model 1",
    "Competitor Brand 2",
    "Competitor Model 2",
]


def product_row(brand="TP-Link", model="EAP-TEST", **overrides):
    values = {
        "Brand": brand,
        "Model": model,
        "Region / HW Version": "US V1",
        "AP Type": "Ceiling",
        "Wi-Fi Standard": "Wi-Fi 7",
        "Supported Wireless Bands": "2.4 / 5 / 6 GHz",
        "Total Spatial Streams": 6,
        "2.4 GHz MIMO": "2×2",
        "5 GHz MIMO": "2×2",
        "6 GHz MIMO": "2×2",
        "2.4 GHz Max Rate (Mbps)": 688,
        "5 GHz Max Rate (Mbps)": 4324,
        "6 GHz Max Rate (Mbps)": 5765,
        "Aggregate Rate (Mbps)": 1,
        "Max Channel Width (MHz)": 320,
        "Ethernet Interfaces": "1× 2.5 GbE RJ45",
        "PoE Input": "802.3at",
        "PoE Output": "No",
        "Max Clients": "300+",
        "IP Rating": "Not Published",
        "Official Source": "https://example.com/product",
        "Last Verified": "2026-07-18",
        "Data Notes": "Test row",
    }
    values.update(overrides)
    return values


def workbook_upload(
    products=None,
    matches=None,
    missing_sheet=None,
    missing_spec_column=None,
    filename="import.xlsx",
):
    products = products or [
        product_row(),
        product_row(brand="Ubiquiti", model="U7-TEST"),
    ]
    matches = matches if matches is not None else [
        ["EAP-TEST", "Ubiquiti", "U7-TEST", "", ""]
    ]
    workbook = Workbook()
    workbook.remove(workbook.active)
    if missing_sheet != "Spec Data":
        spec_sheet = workbook.create_sheet("Spec Data")
        headers = [column for column in SPEC_COLUMNS if column != missing_spec_column]
        spec_sheet.append(headers)
        for product in products:
            spec_sheet.append([product.get(column) for column in headers])
    if missing_sheet != "Match Map":
        match_sheet = workbook.create_sheet("Match Map")
        match_sheet.append(MATCH_COLUMNS)
        for match in matches:
            match_sheet.append(match)
    if missing_sheet != "Field Definitions":
        definitions = workbook.create_sheet("Field Definitions")
        definitions.append(["Field", "Type / Unit", "Definition and Collection Rule"])
        definitions.append(["Supported Wireless Bands", "Text", "Test rule"])
    output = BytesIO()
    workbook.save(output)
    workbook.close()
    return SimpleUploadedFile(
        filename,
        output.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


def product_type_workbook_upload(
    product_type,
    products=None,
    metadata_product_type_code=None,
    filename="product-type-import.xlsx",
):
    profile = get_template_profile(product_type.category.slug, product_type.code)
    products = products or [
        {
            "brand": "TP-Link",
            "model": "TEST-MODEL",
            "region": "Global",
            "hardware_version": "V1",
            "lifecycle_status": "active",
            "official_url": "https://example.com/product",
            "last_verified": "2026-07-27",
        }
    ]
    workbook = Workbook()
    workbook.remove(workbook.active)
    metadata = workbook.create_sheet("Template Metadata")
    metadata.append(["schema_version", "2"])
    metadata.append(["category_slug", profile.category_slug])
    metadata.append(
        [
            "product_type_code",
            metadata_product_type_code or profile.product_type_code,
        ]
    )
    product_sheet = workbook.create_sheet("Product Data")
    headers = [field["code"] for field in profile.fields]
    product_sheet.append(headers)
    for product in products:
        product_sheet.append([product.get(header) for header in headers])
    match_sheet = workbook.create_sheet("Match Map")
    match_sheet.append(
        [
            "tp_link_model",
            "tp_link_region",
            "tp_link_hardware_version",
            "competitor_brand",
            "competitor_model",
            "competitor_region",
            "competitor_hardware_version",
            "match_level",
            "match_reason",
            "source_url",
        ]
    )
    definitions = workbook.create_sheet("Field Definitions")
    definitions.append(
        ["field_code", "中文字段", "字段组", "数据类型", "单位", "必填", "说明"]
    )
    output = BytesIO()
    workbook.save(output)
    workbook.close()
    return SimpleUploadedFile(
        filename,
        output.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


def create_job(
    user,
    upload,
    mode=ImportJob.Mode.PREVIEW,
    product_type=None,
):
    return ImportJob.objects.create(
        uploaded_file=upload,
        uploaded_by=user,
        mode=mode,
        product_type=product_type,
    )
