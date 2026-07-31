import csv
import io
import logging
from dataclasses import dataclass, field
from datetime import date, datetime
from decimal import Decimal, InvalidOperation

from django.core.exceptions import ValidationError
from django.core.files.base import ContentFile
from django.core.validators import URLValidator
from django.db import transaction
from django.utils import timezone
from django.utils.dateparse import parse_date
from openpyxl import load_workbook

from catalog.models import (
    Brand,
    Category,
    ComparisonTemplate,
    Product,
    ProductSpec,
    ProductType,
    SpecDefinition,
    normalize_model_key,
)
from catalog.product_types import infer_product_type_code, product_type_code
from comparison.models import BenchmarkCase, ProductMatch

from .models import ImportJob
from .template_profiles import BASE_FIELDS, SCHEMA_VERSION, get_template_profile

logger = logging.getLogger(__name__)

SPEC_SHEET = "Spec Data"
MATCH_SHEET = "Match Map"
FIELD_DEFINITIONS_SHEET = "Field Definitions"
PRODUCT_SHEET = "Product Data"
METADATA_SHEET = "Template Metadata"

SPEC_COLUMNS = (
    "Brand",
    "Model",
    "Region / HW Version",
    "AP Type",
    "Wi-Fi Standard",
    "Supported Wireless Bands",
    "Total Spatial Streams",
    "2.4 GHz MIMO",
    "5 GHz MIMO",
    "6 GHz MIMO",
    "2.4 GHz Max Rate (Mbps)",
    "5 GHz Max Rate (Mbps)",
    "6 GHz Max Rate (Mbps)",
    "Aggregate Rate (Mbps)",
    "Max Channel Width (MHz)",
    "Ethernet Interfaces",
    "PoE Input",
    "PoE Output",
    "Max Clients",
    "IP Rating",
    "Official Source",
    "Last Verified",
    "Data Notes",
)

MATCH_COLUMNS = (
    "TP-Link Model",
    "Competitor Brand 1",
    "Competitor Model 1",
    "Competitor Brand 2",
    "Competitor Model 2",
)

AP_TYPE_MAP = {
    "ceiling": Product.APType.CEILING,
    "wall": Product.APType.WALL,
    "wall plate": Product.APType.WALL_PLATE,
    "outdoor": Product.APType.OUTDOOR,
    "other": Product.APType.OTHER,
}

SPEC_FIELD_MAP = (
    ("supported_bands", "Supported Wireless Bands", "text"),
    ("total_spatial_streams", "Total Spatial Streams", "number"),
    ("mimo_2g", "2.4 GHz MIMO", "text"),
    ("mimo_5g", "5 GHz MIMO", "text"),
    ("mimo_6g", "6 GHz MIMO", "text"),
    ("rate_2g_mbps", "2.4 GHz Max Rate (Mbps)", "number"),
    ("rate_5g_mbps", "5 GHz Max Rate (Mbps)", "number"),
    ("rate_6g_mbps", "6 GHz Max Rate (Mbps)", "number"),
    ("max_channel_width_mhz", "Max Channel Width (MHz)", "number"),
    ("ethernet_interfaces", "Ethernet Interfaces", "text"),
    ("poe_input", "PoE Input", "text"),
    ("poe_output", "PoE Output", "text"),
    ("max_clients", "Max Clients", "number"),
    ("ip_rating", "IP Rating", "text"),
)


class ImportValidationError(Exception):
    pass


@dataclass
class ImportIssue:
    sheet_name: str
    row_number: int
    column_name: str
    original_value: object
    error_code: str
    message: str

    def as_dict(self):
        return {
            "sheet_name": self.sheet_name,
            "row_number": self.row_number,
            "column_name": self.column_name,
            "original_value": _display(self.original_value),
            "error_code": self.error_code,
            "message": self.message,
        }


@dataclass
class ProductRow:
    row_number: int
    brand: Brand
    model: str
    model_key: str
    region: str
    hardware_version: str
    ap_type: str
    wifi_standard: str
    official_url: str
    verified_date: date | None
    notes: str
    specs: dict
    aggregate_rate: Decimal
    existing_id: int | None = None
    category: Category | None = None
    product_type: ProductType | None = None
    lifecycle_status: str = Product.LifecycleStatus.UNKNOWN
    datasheet_url: str = ""
    launch_date: date | None = None

    @property
    def identity(self):
        return (self.brand_id, self.model_key, self.region, self.hardware_version)

    @property
    def brand_id(self):
        return self.brand.pk

    def preview(self, mode):
        if not self.existing_id:
            action = "create"
        elif mode == ImportJob.Mode.CREATE_UPDATE:
            action = "update"
        else:
            action = "existing"
        return {
            "row": self.row_number,
            "brand": self.brand.name,
            "model": self.model,
            "region": self.region,
            "hardware_version": self.hardware_version,
            "aggregate_rate_mbps": _display(self.aggregate_rate),
            "action": action,
        }


@dataclass
class MatchRow:
    row_number: int
    our_identity: tuple
    competitor_identity: tuple
    region: str
    our_label: str
    competitor_label: str
    existing_id: int | None = None
    match_level: str = ProductMatch.MatchLevel.CORE
    reason: str = ""

    def preview(self, mode):
        if not self.existing_id:
            action = "create"
        elif mode == ImportJob.Mode.CREATE_UPDATE:
            action = "update"
        else:
            action = "existing"
        return {
            "row": self.row_number,
            "our_product": self.our_label,
            "competitor_product": self.competitor_label,
            "region": self.region,
            "action": action,
        }


@dataclass
class ImportPlan:
    products: list[ProductRow] = field(default_factory=list)
    matches: list[MatchRow] = field(default_factory=list)
    issues: list[ImportIssue] = field(default_factory=list)
    total_rows: int = 0

    @property
    def error_rows(self):
        return len({(issue.sheet_name, issue.row_number) for issue in self.issues})

    @property
    def valid_rows(self):
        return max(self.total_rows - self.error_rows, 0)


def _display(value):
    if value is None:
        return ""
    if isinstance(value, (date, datetime)):
        return value.isoformat()
    if isinstance(value, Decimal):
        return format(value, "f").rstrip("0").rstrip(".") or "0"
    return str(value)


def _text(value):
    return _display(value).strip()


def _region_and_version(value):
    parts = _text(value).split(maxsplit=1)
    if not parts:
        return "", ""
    return parts[0], parts[1] if len(parts) == 2 else ""


def _number(plan, value, row_number, column_name, integer=False, required=False):
    if value is None or _text(value) == "":
        if required:
            plan.issues.append(
                ImportIssue(
                    SPEC_SHEET,
                    row_number,
                    column_name,
                    value,
                    "required_value",
                    f"{column_name} is required.",
                )
            )
        return None
    try:
        normalized_number = str(value).strip().replace(",", "")
        if normalized_number.endswith("+"):
            normalized_number = normalized_number[:-1].strip()
        number = Decimal(normalized_number)
    except (InvalidOperation, ValueError):
        plan.issues.append(
            ImportIssue(
                SPEC_SHEET,
                row_number,
                column_name,
                value,
                "invalid_number",
                f"{column_name} must be a non-negative number.",
            )
        )
        return None
    if number < 0 or (integer and number != number.to_integral_value()):
        plan.issues.append(
            ImportIssue(
                SPEC_SHEET,
                row_number,
                column_name,
                value,
                "invalid_number",
                f"{column_name} must be a non-negative integer." if integer else f"{column_name} must be non-negative.",
            )
        )
        return None
    return number


def _verified_date(plan, value, row_number):
    if value is None or _text(value) == "":
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    parsed = parse_date(_text(value))
    if parsed is None:
        plan.issues.append(
            ImportIssue(
                SPEC_SHEET,
                row_number,
                "Last Verified",
                value,
                "invalid_date",
                "Last Verified must be a valid ISO date.",
            )
        )
    return parsed


def _header_map(sheet, required_columns, plan):
    values = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), ())
    headers = {_text(value): index for index, value in enumerate(values) if _text(value)}
    for column in required_columns:
        if column not in headers:
            plan.issues.append(
                ImportIssue(
                    sheet.title,
                    1,
                    column,
                    "",
                    "missing_column",
                    f"Required column '{column}' is missing.",
                )
            )
    return headers


def _row_value(row, headers, column):
    index = headers.get(column)
    return row[index] if index is not None and index < len(row) else None


def _product_ref(product):
    return (product.brand_id, product.model_key, product.region, product.hardware_version)


def _resolve_product(plan, imported_by_model, brand, model, row_number, column_name):
    model_key = normalize_model_key(model)
    candidates = imported_by_model.get((brand.pk, model_key), [])
    if not candidates:
        candidates = list(
            Product.objects.filter(brand=brand, model_key=model_key).select_related("brand")
        )
    if not candidates:
        plan.issues.append(
            ImportIssue(
                MATCH_SHEET,
                row_number,
                column_name,
                model,
                "unknown_product",
                f"Referenced product {brand.name} {model} does not exist.",
            )
        )
        return None
    if len(candidates) > 1:
        plan.issues.append(
            ImportIssue(
                MATCH_SHEET,
                row_number,
                column_name,
                model,
                "ambiguous_product",
                f"Referenced product {brand.name} {model} has multiple versions.",
            )
        )
        return None
    candidate = candidates[0]
    if isinstance(candidate, ProductRow):
        return candidate.identity, candidate.region, f"{brand.name} {candidate.model}"
    return _product_ref(candidate), candidate.region, f"{brand.name} {candidate.model}"


def _modern_date(plan, value, row_number, column_name):
    if value is None or _text(value) == "":
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    parsed = parse_date(_text(value))
    if parsed is None:
        plan.issues.append(
            ImportIssue(
                PRODUCT_SHEET,
                row_number,
                column_name,
                value,
                "invalid_date",
                f"{column_name} must use YYYY-MM-DD.",
            )
        )
    return parsed


def _modern_spec_value(plan, definition, value, row_number):
    raw_value = _display(value)
    text_value = _text(value)
    empty = {
        "value_text": "",
        "value_number": None,
        "value_boolean": None,
        "raw_value": raw_value,
        "value_status": ProductSpec.ValueStatus.PUBLISHED,
    }
    if not text_value:
        return empty
    status_map = {
        "not published": ProductSpec.ValueStatus.NOT_PUBLISHED,
        "not_published": ProductSpec.ValueStatus.NOT_PUBLISHED,
        "not applicable": ProductSpec.ValueStatus.NOT_APPLICABLE,
        "not_applicable": ProductSpec.ValueStatus.NOT_APPLICABLE,
        "unknown": ProductSpec.ValueStatus.UNKNOWN,
    }
    if text_value.casefold() in status_map:
        return {**empty, "value_status": status_map[text_value.casefold()]}
    if definition.data_type in {
        SpecDefinition.DataType.INTEGER,
        SpecDefinition.DataType.DECIMAL,
    }:
        try:
            normalized = text_value.replace(",", "").rstrip("+").strip()
            number = Decimal(normalized)
            if number < 0:
                raise InvalidOperation
            if (
                definition.data_type == SpecDefinition.DataType.INTEGER
                and number != number.to_integral_value()
            ):
                raise InvalidOperation
        except (InvalidOperation, ValueError):
            plan.issues.append(
                ImportIssue(
                    PRODUCT_SHEET,
                    row_number,
                    definition.code,
                    value,
                    "invalid_number",
                    f"{definition.code} must be a non-negative "
                    f"{'integer' if definition.data_type == SpecDefinition.DataType.INTEGER else 'number'}.",
                )
            )
            return empty
        return {**empty, "value_number": number}
    if definition.data_type == SpecDefinition.DataType.BOOLEAN:
        boolean_map = {
            "yes": True,
            "true": True,
            "1": True,
            "no": False,
            "false": False,
            "0": False,
        }
        if text_value.casefold() not in boolean_map:
            plan.issues.append(
                ImportIssue(
                    PRODUCT_SHEET,
                    row_number,
                    definition.code,
                    value,
                    "invalid_boolean",
                    f"{definition.code} must be yes, no, or unknown.",
                )
            )
            return empty
        return {**empty, "value_boolean": boolean_map[text_value.casefold()]}
    return {**empty, "value_text": text_value}


def _resolve_modern_product(
    plan,
    imported_by_identity,
    brand,
    model,
    region,
    hardware_version,
    row_number,
    column_name,
):
    identity = (
        brand.pk,
        normalize_model_key(model),
        region,
        hardware_version,
    )
    imported = imported_by_identity.get(identity)
    if imported:
        return identity, f"{brand.name} {imported.model}"
    product = Product.objects.filter(
        brand=brand,
        model_key=identity[1],
        region=region,
        hardware_version=hardware_version,
    ).first()
    if product:
        return identity, f"{brand.name} {product.model}"
    plan.issues.append(
        ImportIssue(
            MATCH_SHEET,
            row_number,
            column_name,
            model,
            "unknown_product",
            f"Referenced product {brand.name} {model} ({region} {hardware_version}) does not exist.",
        )
    )
    return None


def _build_modern_import_plan(job, workbook):
    plan = ImportPlan()
    if job.product_type_id is None:
        plan.issues.append(
            ImportIssue(
                "Workbook",
                0,
                "Product Type",
                "",
                "missing_product_type",
                "Select a product type before uploading this template.",
            )
        )
        return plan
    product_type = ProductType.objects.select_related("category").get(
        pk=job.product_type_id
    )
    profile = get_template_profile(product_type.category.slug, product_type.code)
    if profile is None:
        plan.issues.append(
            ImportIssue(
                "Workbook",
                0,
                "Product Type",
                product_type.code,
                "unsupported_product_type",
                "No import profile is configured for this product type.",
            )
        )
        return plan

    for sheet_name in (
        METADATA_SHEET,
        PRODUCT_SHEET,
        MATCH_SHEET,
        FIELD_DEFINITIONS_SHEET,
    ):
        if sheet_name not in workbook.sheetnames:
            plan.issues.append(
                ImportIssue(
                    sheet_name,
                    0,
                    "Sheet Name",
                    sheet_name,
                    "missing_sheet",
                    f"Required worksheet '{sheet_name}' is missing.",
                )
            )
    if PRODUCT_SHEET not in workbook.sheetnames:
        return plan

    if METADATA_SHEET in workbook.sheetnames:
        metadata = {
            _text(row[0]): _text(row[1])
            for row in workbook[METADATA_SHEET].iter_rows(
                min_row=1, max_col=2, values_only=True
            )
            if row and _text(row[0])
        }
        expected = {
            "schema_version": SCHEMA_VERSION,
            "category_slug": profile.category_slug,
            "product_type_code": profile.product_type_code,
        }
        for key, expected_value in expected.items():
            actual = metadata.get(key, "")
            if actual != expected_value:
                plan.issues.append(
                    ImportIssue(
                        METADATA_SHEET,
                        0,
                        key,
                        actual,
                        "template_mismatch",
                        f"Template metadata {key} must be '{expected_value}'.",
                    )
                )

    definitions = {
        definition.code: definition
        for definition in SpecDefinition.objects.filter(
            code__in=[
                field["code"]
                for field in profile.fields
                if field["code"] not in {item["code"] for item in BASE_FIELDS}
            ]
        )
    }
    required_columns = tuple(field["code"] for field in profile.fields)
    sheet = workbook[PRODUCT_SHEET]
    headers = _header_map(sheet, required_columns, plan)
    if any(
        issue.error_code == "missing_column"
        and issue.sheet_name == PRODUCT_SHEET
        for issue in plan.issues
    ):
        return plan

    brands = {brand.name.casefold(): brand for brand in Brand.objects.filter(active=True)}
    seen_products = set()
    imported_by_identity = {}
    lifecycle_values = {choice for choice, _label in Product.LifecycleStatus.choices}
    url_validator = URLValidator()
    ap_types = {choice for choice, _label in Product.APType.choices}
    spec_codes = tuple(definitions)

    for row_number, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        brand_name = _text(_row_value(row, headers, "brand"))
        model = _text(_row_value(row, headers, "model"))
        if not brand_name and not model:
            continue
        plan.total_rows += 1
        if not brand_name:
            plan.issues.append(
                ImportIssue(PRODUCT_SHEET, row_number, "brand", "", "required_value", "brand is required.")
            )
            continue
        if not model:
            plan.issues.append(
                ImportIssue(PRODUCT_SHEET, row_number, "model", "", "required_value", "model is required.")
            )
            continue
        brand = brands.get(brand_name.casefold())
        if brand is None:
            plan.issues.append(
                ImportIssue(
                    PRODUCT_SHEET,
                    row_number,
                    "brand",
                    brand_name,
                    "unknown_brand",
                    f"Brand '{brand_name}' is not configured.",
                )
            )
            continue
        region = _text(_row_value(row, headers, "region"))
        hardware_version = _text(_row_value(row, headers, "hardware_version"))
        if not region:
            plan.issues.append(
                ImportIssue(PRODUCT_SHEET, row_number, "region", "", "required_value", "region is required.")
            )
        identity = (brand.pk, normalize_model_key(model), region, hardware_version)
        if identity in seen_products:
            plan.issues.append(
                ImportIssue(
                    PRODUCT_SHEET,
                    row_number,
                    "model",
                    model,
                    "duplicate_product",
                    "The same brand, model, region, and hardware version appears more than once.",
                )
            )
            continue
        seen_products.add(identity)

        lifecycle_status = (
            _text(_row_value(row, headers, "lifecycle_status"))
            or Product.LifecycleStatus.UNKNOWN
        )
        if lifecycle_status not in lifecycle_values:
            plan.issues.append(
                ImportIssue(
                    PRODUCT_SHEET,
                    row_number,
                    "lifecycle_status",
                    lifecycle_status,
                    "invalid_lifecycle_status",
                    f"lifecycle_status must be one of: {', '.join(sorted(lifecycle_values))}.",
                )
            )
            lifecycle_status = Product.LifecycleStatus.UNKNOWN
        official_url = _text(_row_value(row, headers, "official_url"))
        datasheet_url = _text(_row_value(row, headers, "datasheet_url"))
        for column_name, url in (
            ("official_url", official_url),
            ("datasheet_url", datasheet_url),
        ):
            if url:
                try:
                    url_validator(url)
                except ValidationError:
                    plan.issues.append(
                        ImportIssue(
                            PRODUCT_SHEET,
                            row_number,
                            column_name,
                            url,
                            "invalid_url",
                            f"{column_name} must be a valid HTTP or HTTPS URL.",
                        )
                    )
        launch_date = _modern_date(
            plan, _row_value(row, headers, "launch_date"), row_number, "launch_date"
        )
        verified_date = _modern_date(
            plan,
            _row_value(row, headers, "last_verified"),
            row_number,
            "last_verified",
        )
        notes = _text(_row_value(row, headers, "data_notes"))
        specs = {
            code: _modern_spec_value(
                plan,
                definitions[code],
                _row_value(row, headers, code),
                row_number,
            )
            for code in spec_codes
        }
        aggregate_rate = sum(
            (
                specs.get(code, {}).get("value_number") or Decimal("0")
                for code in ("rate_2g_mbps", "rate_5g_mbps", "rate_6g_mbps")
            ),
            Decimal("0"),
        )
        existing = Product.objects.filter(
            brand=brand,
            model_key=identity[1],
            region=region,
            hardware_version=hardware_version,
        ).first()
        if existing and existing.category_id != product_type.category_id:
            plan.issues.append(
                ImportIssue(
                    PRODUCT_SHEET,
                    row_number,
                    "model",
                    model,
                    "category_conflict",
                    f"Existing product belongs to {existing.category.name}; category changes are not allowed by import.",
                )
            )
        if existing and job.mode == ImportJob.Mode.CREATE_ONLY:
            plan.issues.append(
                ImportIssue(
                    PRODUCT_SHEET,
                    row_number,
                    "model",
                    model,
                    "product_exists",
                    "Product already exists and Create only mode was selected.",
                )
            )
        ap_type = (
            product_type.code
            if product_type.category.slug == "access-point"
            and product_type.code in ap_types
            else Product.APType.OTHER
            if product_type.category.slug == "access-point"
            else ""
        )
        product_row = ProductRow(
            row_number=row_number,
            brand=brand,
            model=model,
            model_key=identity[1],
            region=region,
            hardware_version=hardware_version,
            ap_type=ap_type,
            wifi_standard="",
            official_url=official_url,
            verified_date=verified_date,
            notes=notes,
            specs=specs,
            aggregate_rate=aggregate_rate,
            existing_id=existing.pk if existing else None,
            category=product_type.category,
            product_type=product_type,
            lifecycle_status=lifecycle_status,
            datasheet_url=datasheet_url,
            launch_date=launch_date,
        )
        plan.products.append(product_row)
        imported_by_identity[identity] = product_row

    if MATCH_SHEET in workbook.sheetnames:
        match_columns = (
            "tp_link_model",
            "tp_link_region",
            "tp_link_hardware_version",
            "competitor_brand",
            "competitor_model",
            "competitor_region",
            "competitor_hardware_version",
            "match_level",
            "match_reason",
            "source_url",
        )
        match_sheet = workbook[MATCH_SHEET]
        match_headers = _header_map(match_sheet, match_columns, plan)
        if not any(
            issue.error_code == "missing_column"
            and issue.sheet_name == MATCH_SHEET
            for issue in plan.issues
        ):
            own_brand = Brand.objects.filter(active=True, is_own_brand=True).first()
            seen_matches = set()
            match_levels = {
                choice for choice, _label in ProductMatch.MatchLevel.choices
            }
            for row_number, row in enumerate(
                match_sheet.iter_rows(min_row=2, values_only=True), start=2
            ):
                own_model = _text(_row_value(row, match_headers, "tp_link_model"))
                competitor_model = _text(
                    _row_value(row, match_headers, "competitor_model")
                )
                if not own_model and not competitor_model:
                    continue
                plan.total_rows += 1
                competitor_brand_name = _text(
                    _row_value(row, match_headers, "competitor_brand")
                )
                competitor_brand = brands.get(competitor_brand_name.casefold())
                if own_brand is None or competitor_brand is None:
                    plan.issues.append(
                        ImportIssue(
                            MATCH_SHEET,
                            row_number,
                            "competitor_brand",
                            competitor_brand_name,
                            "unknown_brand",
                            "Own brand or competitor brand is not configured.",
                        )
                    )
                    continue
                own_region = _text(
                    _row_value(row, match_headers, "tp_link_region")
                )
                own_hw = _text(
                    _row_value(row, match_headers, "tp_link_hardware_version")
                )
                competitor_region = _text(
                    _row_value(row, match_headers, "competitor_region")
                )
                competitor_hw = _text(
                    _row_value(row, match_headers, "competitor_hardware_version")
                )
                own_ref = _resolve_modern_product(
                    plan,
                    imported_by_identity,
                    own_brand,
                    own_model,
                    own_region,
                    own_hw,
                    row_number,
                    "tp_link_model",
                )
                competitor_ref = _resolve_modern_product(
                    plan,
                    imported_by_identity,
                    competitor_brand,
                    competitor_model,
                    competitor_region,
                    competitor_hw,
                    row_number,
                    "competitor_model",
                )
                if not own_ref or not competitor_ref:
                    continue
                match_identity = (own_ref[0], competitor_ref[0], own_region)
                if match_identity in seen_matches:
                    plan.issues.append(
                        ImportIssue(
                            MATCH_SHEET,
                            row_number,
                            "competitor_model",
                            competitor_model,
                            "duplicate_match",
                            "The same product match appears more than once.",
                        )
                    )
                    continue
                seen_matches.add(match_identity)
                level = (
                    _text(_row_value(row, match_headers, "match_level"))
                    or ProductMatch.MatchLevel.CORE
                )
                if level not in match_levels:
                    plan.issues.append(
                        ImportIssue(
                            MATCH_SHEET,
                            row_number,
                            "match_level",
                            level,
                            "invalid_match_level",
                            f"match_level must be one of: {', '.join(sorted(match_levels))}.",
                        )
                    )
                    level = ProductMatch.MatchLevel.CORE
                existing_match = ProductMatch.objects.filter(
                    our_product__brand_id=own_ref[0][0],
                    our_product__model_key=own_ref[0][1],
                    our_product__region=own_ref[0][2],
                    our_product__hardware_version=own_ref[0][3],
                    competitor_product__brand_id=competitor_ref[0][0],
                    competitor_product__model_key=competitor_ref[0][1],
                    competitor_product__region=competitor_ref[0][2],
                    competitor_product__hardware_version=competitor_ref[0][3],
                    region=own_region,
                ).first()
                if existing_match and job.mode == ImportJob.Mode.CREATE_ONLY:
                    plan.issues.append(
                        ImportIssue(
                            MATCH_SHEET,
                            row_number,
                            "competitor_model",
                            competitor_model,
                            "match_exists",
                            "Product match already exists and Create only mode was selected.",
                        )
                    )
                plan.matches.append(
                    MatchRow(
                        row_number=row_number,
                        our_identity=own_ref[0],
                        competitor_identity=competitor_ref[0],
                        region=own_region,
                        our_label=own_ref[1],
                        competitor_label=competitor_ref[1],
                        existing_id=existing_match.pk if existing_match else None,
                        match_level=level,
                        reason=_text(
                            _row_value(row, match_headers, "match_reason")
                        ),
                    )
                )
    return plan


def build_import_plan(job):
    plan = ImportPlan()
    try:
        workbook = load_workbook(job.uploaded_file.path, read_only=True, data_only=True)
    except Exception as exc:
        plan.issues.append(
            ImportIssue("Workbook", 0, "File", job.uploaded_file.name, "invalid_workbook", str(exc))
        )
        return plan

    if PRODUCT_SHEET in workbook.sheetnames:
        modern_plan = _build_modern_import_plan(job, workbook)
        workbook.close()
        return modern_plan

    missing_sheets = [
        sheet_name
        for sheet_name in (SPEC_SHEET, MATCH_SHEET, FIELD_DEFINITIONS_SHEET)
        if sheet_name not in workbook.sheetnames
    ]
    for sheet_name in missing_sheets:
        plan.issues.append(
            ImportIssue(
                sheet_name,
                0,
                "Sheet Name",
                sheet_name,
                "missing_sheet",
                f"Required worksheet '{sheet_name}' is missing.",
            )
        )
    if SPEC_SHEET not in workbook.sheetnames:
        workbook.close()
        return plan

    brands = {brand.name.casefold(): brand for brand in Brand.objects.filter(active=True)}
    spec_sheet = workbook[SPEC_SHEET]
    spec_headers = _header_map(spec_sheet, SPEC_COLUMNS, plan)
    if any(issue.error_code == "missing_column" and issue.sheet_name == SPEC_SHEET for issue in plan.issues):
        workbook.close()
        return plan

    seen_products = set()
    imported_by_model = {}
    for row_number, row in enumerate(spec_sheet.iter_rows(min_row=2, values_only=True), start=2):
        brand_name = _text(_row_value(row, spec_headers, "Brand"))
        model = _text(_row_value(row, spec_headers, "Model"))
        if not brand_name and not model:
            continue
        plan.total_rows += 1
        if not brand_name:
            plan.issues.append(ImportIssue(SPEC_SHEET, row_number, "Brand", "", "required_value", "Brand is required."))
            continue
        if not model:
            plan.issues.append(ImportIssue(SPEC_SHEET, row_number, "Model", "", "required_value", "Model is required."))
            continue
        brand = brands.get(brand_name.casefold())
        if brand is None:
            plan.issues.append(ImportIssue(SPEC_SHEET, row_number, "Brand", brand_name, "unknown_brand", f"Brand '{brand_name}' is not configured."))
            continue

        region_value = _row_value(row, spec_headers, "Region / HW Version")
        region, hardware_version = _region_and_version(region_value)
        if not region:
            plan.issues.append(ImportIssue(SPEC_SHEET, row_number, "Region / HW Version", region_value, "required_value", "Region / HW Version is required."))

        model_key = normalize_model_key(model)
        identity = (brand.pk, model_key, region, hardware_version)
        if identity in seen_products:
            plan.issues.append(ImportIssue(SPEC_SHEET, row_number, "Model", model, "duplicate_product", "The same brand, model, region, and hardware version appears more than once."))
            continue
        seen_products.add(identity)

        ap_type_value = _text(_row_value(row, spec_headers, "AP Type"))
        ap_type = AP_TYPE_MAP.get(ap_type_value.casefold())
        if ap_type is None:
            plan.issues.append(ImportIssue(SPEC_SHEET, row_number, "AP Type", ap_type_value, "invalid_ap_type", f"AP Type '{ap_type_value}' is not allowed."))
            ap_type = Product.APType.OTHER

        wifi_standard = _text(_row_value(row, spec_headers, "Wi-Fi Standard"))
        if wifi_standard != "Wi-Fi 7":
            plan.issues.append(ImportIssue(SPEC_SHEET, row_number, "Wi-Fi Standard", wifi_standard, "invalid_wifi_standard", "Only Wi-Fi 7 is supported in this phase."))

        official_url = _text(_row_value(row, spec_headers, "Official Source"))
        if official_url:
            try:
                URLValidator()(official_url)
            except ValidationError:
                plan.issues.append(ImportIssue(SPEC_SHEET, row_number, "Official Source", official_url, "invalid_url", "Official Source must be a valid HTTP or HTTPS URL."))

        verified_date = _verified_date(plan, _row_value(row, spec_headers, "Last Verified"), row_number)
        notes = _text(_row_value(row, spec_headers, "Data Notes"))
        specs = {}
        numeric_columns = {
            "Total Spatial Streams": True,
            "Max Clients": True,
            "2.4 GHz Max Rate (Mbps)": False,
            "5 GHz Max Rate (Mbps)": False,
            "6 GHz Max Rate (Mbps)": False,
            "Max Channel Width (MHz)": False,
        }
        for code, column, value_type in SPEC_FIELD_MAP:
            original = _row_value(row, spec_headers, column)
            if value_type == "number":
                parsed = _number(
                    plan,
                    original,
                    row_number,
                    column,
                    integer=numeric_columns[column],
                    required=column == "Total Spatial Streams",
                )
                specs[code] = {"value_number": parsed, "value_text": "", "raw_value": _display(original)}
            else:
                specs[code] = {"value_number": None, "value_text": _text(original), "raw_value": _display(original)}

        bands = specs["supported_bands"]["value_text"]
        rate_2g = specs["rate_2g_mbps"]["value_number"]
        rate_5g = specs["rate_5g_mbps"]["value_number"]
        rate_6g = specs["rate_6g_mbps"]["value_number"]
        if bands not in {"2.4 / 5 GHz", "2.4 / 5 / 6 GHz"}:
            plan.issues.append(ImportIssue(SPEC_SHEET, row_number, "Supported Wireless Bands", bands, "invalid_bands", "Supported bands must be '2.4 / 5 GHz' or '2.4 / 5 / 6 GHz'."))
        if rate_2g is None or rate_5g is None:
            plan.issues.append(ImportIssue(SPEC_SHEET, row_number, "Supported Wireless Bands", bands, "band_rate_conflict", "A product supporting 2.4 / 5 GHz must have both corresponding rates."))
        if bands == "2.4 / 5 GHz" and rate_6g is not None:
            plan.issues.append(ImportIssue(SPEC_SHEET, row_number, "6 GHz Max Rate (Mbps)", rate_6g, "band_rate_conflict", "A dual-band product must not include a 6 GHz rate."))
        if bands == "2.4 / 5 / 6 GHz":
            if rate_6g is None:
                plan.issues.append(ImportIssue(SPEC_SHEET, row_number, "6 GHz Max Rate (Mbps)", rate_6g, "band_rate_conflict", "A tri-band product must include a 6 GHz rate."))
            missing_mimo = [column for code, column in (("mimo_2g", "2.4 GHz MIMO"), ("mimo_5g", "5 GHz MIMO"), ("mimo_6g", "6 GHz MIMO")) if not specs[code]["value_text"]]
            for column in missing_mimo:
                plan.issues.append(ImportIssue(SPEC_SHEET, row_number, column, "", "missing_mimo", "Tri-band products must provide all three MIMO fields."))

        aggregate_rate = sum((value or Decimal("0") for value in (rate_2g, rate_5g, rate_6g)), Decimal("0"))
        existing = Product.objects.filter(
            brand=brand,
            model_key=model_key,
            region=region,
            hardware_version=hardware_version,
        ).first()
        if existing and job.mode == ImportJob.Mode.CREATE_ONLY:
            plan.issues.append(ImportIssue(SPEC_SHEET, row_number, "Model", model, "product_exists", "Product already exists and Create only mode was selected."))

        product_row = ProductRow(
            row_number=row_number,
            brand=brand,
            model=model,
            model_key=model_key,
            region=region,
            hardware_version=hardware_version,
            ap_type=ap_type,
            wifi_standard=wifi_standard,
            official_url=official_url,
            verified_date=verified_date,
            notes=notes,
            specs=specs,
            aggregate_rate=aggregate_rate,
            existing_id=existing.pk if existing else None,
        )
        plan.products.append(product_row)
        imported_by_model.setdefault((brand.pk, model_key), []).append(product_row)

    if MATCH_SHEET in workbook.sheetnames:
        match_sheet = workbook[MATCH_SHEET]
        match_headers = _header_map(match_sheet, MATCH_COLUMNS, plan)
        if not any(issue.error_code == "missing_column" and issue.sheet_name == MATCH_SHEET for issue in plan.issues):
            seen_matches = set()
            own_brand = next((brand for brand in brands.values() if brand.is_own_brand), None)
            for row_number, row in enumerate(match_sheet.iter_rows(min_row=2, values_only=True), start=2):
                own_model = _text(_row_value(row, match_headers, "TP-Link Model"))
                pairs = (
                    (_text(_row_value(row, match_headers, "Competitor Brand 1")), _text(_row_value(row, match_headers, "Competitor Model 1")), "Competitor Model 1"),
                    (_text(_row_value(row, match_headers, "Competitor Brand 2")), _text(_row_value(row, match_headers, "Competitor Model 2")), "Competitor Model 2"),
                )
                for competitor_brand_name, competitor_model, model_column in pairs:
                    if not competitor_brand_name and not competitor_model:
                        continue
                    plan.total_rows += 1
                    if not own_model:
                        plan.issues.append(ImportIssue(MATCH_SHEET, row_number, "TP-Link Model", own_model, "required_value", "TP-Link Model is required."))
                        continue
                    if not competitor_brand_name or not competitor_model:
                        plan.issues.append(ImportIssue(MATCH_SHEET, row_number, model_column, competitor_model, "incomplete_match", "Competitor brand and model must both be provided."))
                        continue
                    competitor_brand = brands.get(competitor_brand_name.casefold())
                    if competitor_brand is None:
                        plan.issues.append(ImportIssue(MATCH_SHEET, row_number, model_column.replace("Model", "Brand"), competitor_brand_name, "unknown_brand", f"Brand '{competitor_brand_name}' is not configured."))
                        continue
                    if competitor_brand.is_own_brand:
                        plan.issues.append(ImportIssue(MATCH_SHEET, row_number, model_column.replace("Model", "Brand"), competitor_brand_name, "invalid_competitor", "TP-Link cannot be imported as a competitor product."))
                        continue
                    if own_brand is None:
                        plan.issues.append(ImportIssue(MATCH_SHEET, row_number, "TP-Link Model", own_model, "missing_own_brand", "No active own brand is configured."))
                        continue
                    own_ref = _resolve_product(plan, imported_by_model, own_brand, own_model, row_number, "TP-Link Model")
                    competitor_ref = _resolve_product(plan, imported_by_model, competitor_brand, competitor_model, row_number, model_column)
                    if not own_ref or not competitor_ref:
                        continue
                    own_identity, region, own_label = own_ref
                    competitor_identity, _, competitor_label = competitor_ref
                    if own_identity == competitor_identity:
                        plan.issues.append(ImportIssue(MATCH_SHEET, row_number, model_column, competitor_model, "self_match", "A product cannot be matched to itself."))
                        continue
                    match_identity = (own_identity, competitor_identity, region)
                    if match_identity in seen_matches:
                        plan.issues.append(ImportIssue(MATCH_SHEET, row_number, model_column, competitor_model, "duplicate_match", "The same product match appears more than once."))
                        continue
                    seen_matches.add(match_identity)
                    existing_match = None
                    if all(isinstance(value, int) for value in (own_identity[0], competitor_identity[0])):
                        own_product = Product.objects.filter(
                            brand_id=own_identity[0], model_key=own_identity[1], region=own_identity[2], hardware_version=own_identity[3]
                        ).first()
                        competitor_product = Product.objects.filter(
                            brand_id=competitor_identity[0], model_key=competitor_identity[1], region=competitor_identity[2], hardware_version=competitor_identity[3]
                        ).first()
                        if own_product and competitor_product:
                            existing_match = ProductMatch.objects.filter(
                                our_product=own_product,
                                competitor_product=competitor_product,
                                region=region,
                            ).first()
                    if existing_match and job.mode == ImportJob.Mode.CREATE_ONLY:
                        plan.issues.append(ImportIssue(MATCH_SHEET, row_number, model_column, competitor_model, "match_exists", "Product match already exists and Create only mode was selected."))
                    plan.matches.append(
                        MatchRow(
                            row_number=row_number,
                            our_identity=own_identity,
                            competitor_identity=competitor_identity,
                            region=region,
                            our_label=own_label,
                            competitor_label=competitor_label,
                            existing_id=existing_match.pk if existing_match else None,
                        )
                    )

    workbook.close()
    return plan


def _audit(summary, event, user, **details):
    audit = list(summary.get("audit", []))
    audit.append(
        {
            "event": event,
            "at": timezone.now().isoformat(),
            "user_id": user.pk if user else None,
            **details,
        }
    )
    summary["audit"] = audit


def _save_error_report(job, issues):
    if job.error_report:
        job.error_report.delete(save=False)
    if not issues:
        job.error_report = ""
        return
    output = io.StringIO(newline="")
    writer = csv.writer(output)
    writer.writerow(("Sheet Name", "Row Number", "Column Name", "Original Value", "Error Code", "Human-readable Message"))
    for issue in issues:
        writer.writerow(
            (
                issue.sheet_name,
                issue.row_number,
                issue.column_name,
                _display(issue.original_value),
                issue.error_code,
                issue.message,
            )
        )
    job.error_report.save(
        f"import-{job.pk}-errors.csv",
        ContentFile(output.getvalue().encode("utf-8-sig")),
        save=False,
    )


def validate_import_job(job):
    job.status = ImportJob.Status.VALIDATING
    job.save(update_fields=("status",))
    plan = build_import_plan(job)
    previous_summary = job.summary or {}
    summary = {
        "product_count": len(plan.products),
        "match_count": len(plan.matches),
        "product_preview": [row.preview(job.mode) for row in plan.products[:10]],
        "match_preview": [row.preview(job.mode) for row in plan.matches[:10]],
        "errors": [issue.as_dict() for issue in plan.issues[:50]],
        "audit": previous_summary.get("audit", []),
    }
    _audit(summary, "validated", job.uploaded_by, errors=len(plan.issues))
    job.total_rows = plan.total_rows
    job.valid_rows = plan.valid_rows
    job.error_rows = plan.error_rows
    job.status = ImportJob.Status.INVALID if plan.issues else ImportJob.Status.READY
    job.summary = summary
    _save_error_report(job, plan.issues)
    job.save()
    logger.info(
        "Validated import job %s: status=%s products=%s matches=%s errors=%s",
        job.pk,
        job.status,
        len(plan.products),
        len(plan.matches),
        len(plan.issues),
    )
    return plan


def execute_import_job(job):
    if job.mode == ImportJob.Mode.PREVIEW:
        raise ImportValidationError("Preview-only jobs cannot be imported.")
    if job.status != ImportJob.Status.READY:
        raise ImportValidationError("Only a ready import job can be confirmed.")
    plan = build_import_plan(job)
    if plan.issues:
        validate_import_job(job)
        raise ImportValidationError("The workbook is no longer valid. Review the error report.")

    counters = {
        "products_created": 0,
        "products_updated": 0,
        "specs_created": 0,
        "specs_updated": 0,
        "matches_created": 0,
        "matches_updated": 0,
    }
    try:
        with transaction.atomic():
            locked_job = ImportJob.objects.select_for_update().get(pk=job.pk)
            if locked_job.status != ImportJob.Status.READY:
                raise ImportValidationError("Import job status changed before confirmation.")
            legacy_category = Category.objects.get(slug="access-point")
            definition_codes = {
                code for row in plan.products for code in row.specs
            }
            definitions = {
                definition.code: definition
                for definition in SpecDefinition.objects.filter(
                    code__in=definition_codes
                )
            }
            missing_definitions = definition_codes - definitions.keys()
            if missing_definitions:
                raise ImportValidationError(f"Missing specification definitions: {', '.join(sorted(missing_definitions))}")

            product_objects = {}
            for row in plan.products:
                category = row.category or legacy_category
                product = Product.objects.filter(
                    brand=row.brand,
                    model_key=row.model_key,
                    region=row.region,
                    hardware_version=row.hardware_version,
                ).first()
                if product is None:
                    product = Product(brand=row.brand, category=category, created_by=job.uploaded_by)
                    counters["products_created"] += 1
                else:
                    if job.mode == ImportJob.Mode.CREATE_ONLY:
                        raise ImportValidationError(f"Product already exists: {row.brand.name} {row.model}")
                    counters["products_updated"] += 1
                product.model = row.model
                product.category = category
                product.product_type = row.product_type or ProductType.objects.filter(
                    category=category,
                    code=infer_product_type_code(
                        category.slug, row.model, row.ap_type
                    ),
                    active=True,
                ).first()
                product.region = row.region
                product.hardware_version = row.hardware_version
                product.ap_type = row.ap_type
                product.wifi_standard = row.wifi_standard
                product.lifecycle_status = row.lifecycle_status
                product.official_url = row.official_url
                product.datasheet_url = row.datasheet_url
                product.launch_date = row.launch_date
                product.notes = row.notes
                product.is_published = True
                product.updated_by = job.uploaded_by
                product.save()
                product_objects[row.identity] = product

                for code, values in row.specs.items():
                    if values["value_number"] is None and not values["value_text"]:
                        continue
                    value_text = (values["value_text"] or "").strip()
                    status_map = {
                        "not published": ProductSpec.ValueStatus.NOT_PUBLISHED,
                        "not applicable": ProductSpec.ValueStatus.NOT_APPLICABLE,
                        "unknown": ProductSpec.ValueStatus.UNKNOWN,
                    }
                    value_status = values.get("value_status") or status_map.get(
                        value_text.casefold(),
                        ProductSpec.ValueStatus.PUBLISHED,
                    )
                    typed_values = {
                        **values,
                        "value_status": value_status,
                        "unit": definitions[code].unit,
                        "normalized_value": (
                            str(values["value_number"])
                            if values["value_number"] is not None
                            else (
                                "true"
                                if values.get("value_boolean") is True
                                else "false"
                                if values.get("value_boolean") is False
                                else value_text
                            )
                        ),
                    }
                    if value_status != ProductSpec.ValueStatus.PUBLISHED:
                        typed_values["value_text"] = ""
                        typed_values["value_number"] = None
                        typed_values["value_boolean"] = None
                        typed_values["normalized_value"] = ""
                    _, created = ProductSpec.objects.update_or_create(
                        product=product,
                        definition=definitions[code],
                        defaults={
                            **typed_values,
                            "source_url": "",
                            "source_note": row.notes,
                            "verified_date": row.verified_date,
                            "updated_by": job.uploaded_by,
                        },
                    )
                    counters["specs_created" if created else "specs_updated"] += 1

            for row in plan.matches:
                our_product = product_objects.get(row.our_identity)
                if our_product is None:
                    our_product = Product.objects.get(
                        brand_id=row.our_identity[0],
                        model_key=row.our_identity[1],
                        region=row.our_identity[2],
                        hardware_version=row.our_identity[3],
                    )
                competitor_product = product_objects.get(row.competitor_identity)
                if competitor_product is None:
                    competitor_product = Product.objects.get(
                        brand_id=row.competitor_identity[0],
                        model_key=row.competitor_identity[1],
                        region=row.competitor_identity[2],
                        hardware_version=row.competitor_identity[3],
                    )
                match = ProductMatch.objects.filter(
                    our_product=our_product,
                    competitor_product=competitor_product,
                    region=row.region,
                ).first()
                template = ComparisonTemplate.objects.filter(
                    category=our_product.category,
                    form_factor__in=(product_type_code(our_product), ""),
                    active=True,
                ).order_by("-form_factor", "-version").first()
                benchmark_case, _ = BenchmarkCase.objects.get_or_create(
                    anchor_product=our_product,
                    region=row.region,
                    name=f"{our_product.model} {row.region} competitor benchmark",
                    defaults={
                        "template": template,
                        "status": BenchmarkCase.Status.APPROVED,
                        "scenario": "Standard product benchmark",
                        "created_by": job.uploaded_by,
                    },
                )
                if match is None:
                    match = ProductMatch(created_by=job.uploaded_by)
                    counters["matches_created"] += 1
                else:
                    if job.mode == ImportJob.Mode.CREATE_ONLY:
                        raise ImportValidationError(f"Product match already exists: {row.our_label} -> {row.competitor_label}")
                    counters["matches_updated"] += 1
                match.our_product = our_product
                match.competitor_product = competitor_product
                match.benchmark_case = benchmark_case
                match.region = row.region
                match.match_type = ProductMatch.MatchType.DIRECT
                match.match_level = row.match_level
                match.status = ProductMatch.Status.CONFIRMED
                match.reason = (
                    row.reason or "Imported from the product-type Excel match map."
                )
                match.updated_by = job.uploaded_by
                match.save()

            summary = dict(locked_job.summary or {})
            summary["result"] = counters
            _audit(summary, "imported", job.uploaded_by, **counters)
            locked_job.summary = summary
            locked_job.status = ImportJob.Status.IMPORTED
            locked_job.imported_at = timezone.now()
            locked_job.save(update_fields=("summary", "status", "imported_at"))
    except Exception as exc:
        job.refresh_from_db()
        summary = dict(job.summary or {})
        _audit(summary, "failed", job.uploaded_by, error=str(exc))
        job.status = ImportJob.Status.FAILED
        job.summary = summary
        job.save(update_fields=("status", "summary"))
        logger.exception("Import job %s failed and was rolled back", job.pk)
        raise

    job.refresh_from_db()
    logger.info("Imported job %s: %s", job.pk, counters)
    return counters
